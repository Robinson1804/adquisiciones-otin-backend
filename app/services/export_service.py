"""Export service — C5: Excel + PDF generation (read-only, in-memory).

All logic lives here; routers only call build_excel / build_pdf and stream
the returned bytes. No DB mutations. No disk I/O.

Design authority: design #162.
"""
from __future__ import annotations

import io
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font
from fastapi import HTTPException, status
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.etapa import EtapaRegistro
from app.models.montos import MontosProceso
from app.models.proceso import Proceso
from app.services.etapas_catalogo import (
    COD_A_FASE,
    ETAPAS_CATALOGO,
    FASES,
    ORDEN_ETAPAS,
    fase_de_cod,
)
from app.services.etapas_service import calcular_progreso


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fmt_date(d) -> str:
    """Return ISO date string or empty string for None."""
    if d is None:
        return ""
    return d.strftime("%Y-%m-%d")


def _fmt_datetime(dt) -> str:
    """Return datetime string or empty string for None."""
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M")


def _fmt_money(v) -> str:
    """Format Decimal/float as S/ #,##0.00 or empty string for None."""
    if v is None:
        return ""
    return f"S/ {float(v):,.2f}"


def _fmt_float(v) -> float | str:
    """Return float or empty string for None."""
    if v is None:
        return ""
    return float(v)


def _derive_fase_label(proceso: Proceso, etapas_rows: list[EtapaRegistro]) -> str:
    """Derive the human-readable fase label for a proceso, mirroring dashboard logic."""
    if proceso.estado == "CULMINADO":
        fase_key = "F5"
    else:
        progreso = calcular_progreso(etapas_rows)
        if progreso.etapa_actual and progreso.etapa_actual in COD_A_FASE:
            fase_key = fase_de_cod(progreso.etapa_actual)
        else:
            fase_key = "F1"
    return f"{fase_key} - {FASES[fase_key]['label']}"


def _get_active_proceso_or_404(db: Session, proceso_id: int) -> Proceso:
    """Return proceso or raise 404 if missing or soft-deleted."""
    p = db.get(Proceso, proceso_id)
    if p is None or p.eliminado_en is not None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Proceso no encontrado",
        )
    return p


# ---------------------------------------------------------------------------
# T1 — build_excel
# ---------------------------------------------------------------------------

def build_excel(db: Session, anno: int) -> bytes:
    """Build a .xlsx workbook for the given year and return its bytes.

    3 sheets: Procesos, Etapas, Montos.
    Empty year → valid workbook with headers only (no 404, per spec E7).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default empty sheet

    # --- Query base ---
    procesos = db.execute(
        select(Proceso).where(
            Proceso.anno == anno,
            Proceso.eliminado_en.is_(None),
        ).order_by(Proceso.id)
    ).scalars().all()

    proceso_ids = [p.id for p in procesos]

    # Fetch all etapas for these processes
    etapas_rows: list[EtapaRegistro] = []
    if proceso_ids:
        etapas_rows = db.execute(
            select(EtapaRegistro).where(
                EtapaRegistro.proceso_id.in_(proceso_ids)
            ).order_by(EtapaRegistro.proceso_id, EtapaRegistro.id)
        ).scalars().all()

    # Fetch montos LEFT JOIN style
    montos_by_proc: dict[int, MontosProceso] = {}
    if proceso_ids:
        montos_list = db.execute(
            select(MontosProceso).where(
                MontosProceso.proceso_id.in_(proceso_ids)
            )
        ).scalars().all()
        montos_by_proc = {m.proceso_id: m for m in montos_list}

    # Group etapas by proceso_id
    etapas_by_proc: dict[int, list[EtapaRegistro]] = {}
    for row in etapas_rows:
        etapas_by_proc.setdefault(row.proceso_id, []).append(row)

    # ---- Sheet: Procesos ----
    ws_proc = wb.create_sheet("Procesos")
    _style_header_row(ws_proc, [
        "id_proceso", "requerimiento", "tipo", "unidad_resp",
        "areas_usuarias", "pim", "estado", "fase_actual",
        "fecha_creacion", "motivo_cancel", "creado_por",
    ])

    for p in procesos:
        rows_etapas = etapas_by_proc.get(p.id, [])
        fase_label = _derive_fase_label(p, rows_etapas)
        areas = "; ".join(p.areas_usuarias) if p.areas_usuarias else ""
        ws_proc.append([
            p.id_proceso,
            p.requerimiento,
            p.tipo or "",
            p.unidad_resp or "",
            areas,
            _fmt_float(p.pim),
            p.estado,
            fase_label,
            _fmt_datetime(p.fecha_creacion),
            p.motivo_cancel or "",
            p.creado_por or "",
        ])

    # ---- Sheet: Etapas ----
    ws_etap = wb.create_sheet("Etapas")
    _style_header_row(ws_etap, [
        "id_proceso", "id_proceso_legible", "codigo_etapa", "nombre_etapa",
        "area_responsable", "area_usuaria", "fecha_inicio", "fecha_fin",
        "dias", "es_bucle", "nro_ronda", "motivo_bucle",
        "estado_etapa", "resultado_eval", "monto_cert",
        "nro_ocs", "monto_ocs", "plazo_entrega", "observaciones",
    ])

    # Map proceso_id → id_proceso string
    id_proceso_map = {p.id: p.id_proceso for p in procesos}

    # Sort by (proceso_id, ORDEN_ETAPAS index, nro_ronda)
    orden_idx = {cod: i for i, cod in enumerate(ORDEN_ETAPAS)}
    sorted_etapas = sorted(
        etapas_rows,
        key=lambda r: (
            r.proceso_id,
            orden_idx.get(r.codigo_etapa, 999),
            r.nro_ronda,
        ),
    )

    for r in sorted_etapas:
        ws_etap.append([
            r.proceso_id,
            id_proceso_map.get(r.proceso_id, ""),
            r.codigo_etapa,
            r.nombre_etapa,
            r.area_responsable or "",
            r.area_usuaria or "",
            _fmt_date(r.fecha_inicio),
            _fmt_date(r.fecha_fin),
            r.dias if r.dias is not None else "",
            "SI" if r.es_bucle else "NO",
            r.nro_ronda,
            r.motivo_bucle or "",
            r.estado_etapa,
            r.resultado_eval or "",
            _fmt_float(r.monto_cert),
            r.nro_ocs or "",
            _fmt_float(r.monto_ocs),
            r.plazo_entrega if r.plazo_entrega is not None else "",
            r.observaciones or "",
        ])

    # ---- Sheet: Montos ----
    ws_mont = wb.create_sheet("Montos")
    _style_header_row(ws_mont, [
        "id_proceso", "id_proceso_legible", "valor_em",
        "monto_cert_total", "nro_ocs", "monto_ocs",
        "plazo_entrega", "fecha_inicio_srv",
    ])

    for p in procesos:
        m = montos_by_proc.get(p.id)
        ws_mont.append([
            p.id,
            p.id_proceso,
            _fmt_float(m.valor_em) if m else "",
            _fmt_float(m.monto_cert_total) if m else "",
            m.nro_ocs if m and m.nro_ocs else "",
            _fmt_float(m.monto_ocs) if m else "",
            m.plazo_entrega if m and m.plazo_entrega is not None else "",
            _fmt_date(m.fecha_inicio_srv) if m else "",
        ])

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _style_header_row(ws, headers: list[str]) -> None:
    """Append a bold header row and freeze pane A2."""
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# T2 — build_pdf
# ---------------------------------------------------------------------------

def build_pdf(db: Session, proceso_id: int) -> bytes:
    """Build a PDF for one proceso and return its bytes.

    Raises 404 if proceso is missing or soft-deleted.
    """
    proceso = _get_active_proceso_or_404(db, proceso_id)

    # Load etapas and montos
    etapas_rows = db.execute(
        select(EtapaRegistro).where(
            EtapaRegistro.proceso_id == proceso_id
        )
    ).scalars().all()

    montos = db.execute(
        select(MontosProceso).where(
            MontosProceso.proceso_id == proceso_id
        )
    ).scalars().first()

    # Pre-compute fase
    fase_label = _derive_fase_label(proceso, list(etapas_rows))

    # Group etapas by code
    etapas_by_cod: dict[str, list[EtapaRegistro]] = {}
    for r in etapas_rows:
        etapas_by_cod.setdefault(r.codigo_etapa, []).append(r)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    story = []

    # ---- 1. Header ----
    story.append(Paragraph("INEI — OTIN | Adquisiciones TIC", styles["Title"]))
    story.append(Paragraph("Resumen Ejecutivo de Proceso de Adquisición", styles["Heading2"]))
    story.append(Paragraph(
        f"<b>ID Proceso:</b> {proceso.id_proceso} &nbsp;&nbsp; "
        f"<b>Estado:</b> {proceso.estado}",
        styles["Normal"],
    ))
    story.append(Spacer(1, 0.4 * cm))

    # ---- 2. Ficha ----
    areas_str = "; ".join(proceso.areas_usuarias) if proceso.areas_usuarias else "—"
    pim_str = _fmt_money(proceso.pim) if proceso.pim else "—"

    ficha_data = [
        ["Campo", "Valor"],
        ["ID Proceso", proceso.id_proceso],
        ["Requerimiento", proceso.requerimiento],
        ["Tipo", proceso.tipo or "—"],
        ["Unidad Responsable", proceso.unidad_resp or "—"],
        ["Áreas Usuarias", areas_str],
        ["PIM", pim_str],
        ["Estado", proceso.estado],
        ["Fase Actual", fase_label],
        ["Fecha Creación", _fmt_datetime(proceso.fecha_creacion)],
    ]

    ficha_table = Table(ficha_data, colWidths=[5 * cm, 12 * cm])
    ficha_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 1), (0, -1), colors.lightgrey),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
    ]))
    story.append(ficha_table)
    story.append(Spacer(1, 0.5 * cm))

    # ---- 3. Stage status table ----
    story.append(Paragraph("Estado de Etapas", styles["Heading3"]))

    stage_header = ["Código", "Nombre", "Área", "Estado", "Días", "Fecha Fin"]
    stage_data = [stage_header]

    for cod in ORDEN_ETAPAS:
        spec = ETAPAS_CATALOGO[cod]
        rows_for_cod = etapas_by_cod.get(cod, [])
        if not rows_for_cod:
            stage_data.append([cod, spec.nombre[:40], spec.area_responsable, "PENDIENTE", "—", "—"])
        else:
            for r in sorted(rows_for_cod, key=lambda x: x.nro_ronda):
                ronda_suffix = f" (r{r.nro_ronda})" if r.es_bucle and r.nro_ronda > 1 else ""
                stage_data.append([
                    cod + ronda_suffix,
                    spec.nombre[:40],
                    r.area_responsable or spec.area_responsable,
                    r.estado_etapa,
                    str(r.dias) if r.dias is not None else "—",
                    _fmt_date(r.fecha_fin) or "—",
                ])

    col_widths = [1.8 * cm, 6.5 * cm, 2.2 * cm, 2.0 * cm, 1.2 * cm, 2.5 * cm]
    stage_table = Table(stage_data, colWidths=col_widths, repeatRows=1)
    stage_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4a4a4a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f0f0")]),
        ("WORDWRAP", (1, 0), (1, -1), True),
    ]))
    story.append(stage_table)
    story.append(Spacer(1, 0.5 * cm))

    # ---- 4. Montos section ----
    story.append(Paragraph("Montos del Proceso", styles["Heading3"]))

    if montos:
        montos_data = [
            ["Concepto", "Valor"],
            ["Valor Estimado de Mercado (EM)", _fmt_money(montos.valor_em) or "—"],
            ["Monto Certificado Total", _fmt_money(montos.monto_cert_total) or "—"],
            ["Número de OCS", montos.nro_ocs or "—"],
            ["Monto OCS", _fmt_money(montos.monto_ocs) or "—"],
            ["Plazo de Entrega (días)", str(montos.plazo_entrega) if montos.plazo_entrega is not None else "—"],
            ["Fecha Inicio Servicio", _fmt_date(montos.fecha_inicio_srv) or "—"],
        ]
        montos_table = Table(montos_data, colWidths=[8 * cm, 9 * cm])
        montos_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 1), (0, -1), colors.lightgrey),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ]))
        story.append(montos_table)
    else:
        story.append(Paragraph("Sin montos registrados.", styles["Normal"]))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()
