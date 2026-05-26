"""Backend tests — C5 export endpoints (Excel + PDF).

Tests: build_excel valid data, empty year, content-type, content-disposition,
       build_pdf valid bytes, PDF endpoint, 404 not found, 404 soft-deleted,
       422 missing anno, 401 unauthenticated, soft-deleted excluded from excel.

Uses client + db_session fixtures from conftest.py.
"""
import io
from datetime import datetime

import openpyxl
import pytest

from app.models.etapa import EtapaRegistro
from app.models.montos import MontosProceso
from app.models.proceso import Proceso
from app.services.export_service import build_excel, build_pdf


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_proceso(
    db_session,
    id_proceso: str = "2026-001",
    anno: int = 2026,
    estado: str = "EN PROCESO",
    eliminado_en=None,
) -> Proceso:
    p = Proceso(
        id_proceso=id_proceso,
        requerimiento=f"Requerimiento test {id_proceso}",
        tipo="BIEN",
        unidad_resp="OTIN",
        areas_usuarias=["DTDIS", "DTSIS"],
        pim=None,
        estado=estado,
        anno=anno,
        creado_por="testsetup",
        eliminado_en=eliminado_en,
    )
    db_session.add(p)
    db_session.flush()
    return p


def _make_etapa(
    db_session,
    proceso_id: int,
    cod: str = "E01",
    estado: str = "COMPLETADO",
) -> EtapaRegistro:
    row = EtapaRegistro(
        proceso_id=proceso_id,
        codigo_etapa=cod,
        nombre_etapa=f"Etapa {cod}",
        area_responsable="OTIN",
        estado_etapa=estado,
        nro_ronda=1,
        registrado_por="testsetup",
    )
    db_session.add(row)
    db_session.flush()
    return row


def _make_montos(db_session, proceso_id: int) -> MontosProceso:
    m = MontosProceso(
        proceso_id=proceso_id,
        valor_em=None,
        monto_cert_total=None,
    )
    db_session.add(m)
    db_session.flush()
    return m


# ---------------------------------------------------------------------------
# T1 — build_excel tests (service-level)
# ---------------------------------------------------------------------------

def test_build_excel_with_data(db_session):
    """2 procesos → valid workbook with 3 sheets, 2 data rows in Procesos."""
    p1 = _make_proceso(db_session, "2026-001")
    p2 = _make_proceso(db_session, "2026-002")
    _make_etapa(db_session, p1.id, "E01", "COMPLETADO")
    _make_etapa(db_session, p2.id, "E02", "PENDIENTE")
    _make_montos(db_session, p1.id)

    data = build_excel(db_session, 2026)

    assert isinstance(data, bytes)
    assert len(data) > 0

    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Procesos", "Etapas", "Montos"]

    ws_proc = wb["Procesos"]
    rows = list(ws_proc.iter_rows(values_only=True))
    assert len(rows) == 3  # 1 header + 2 data

    ws_etap = wb["Etapas"]
    etapa_rows = list(ws_etap.iter_rows(values_only=True))
    assert len(etapa_rows) == 3  # 1 header + 2 etapa rows

    ws_mont = wb["Montos"]
    montos_rows = list(ws_mont.iter_rows(values_only=True))
    assert len(montos_rows) == 3  # 1 header + 2 proceso rows


def test_build_excel_empty_year(db_session):
    """No procesos in year 2099 → valid workbook with header-only sheets (no 404)."""
    data = build_excel(db_session, 2099)

    assert isinstance(data, bytes)
    assert len(data) > 0

    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Procesos", "Etapas", "Montos"]

    for sheet_name in ["Procesos", "Etapas", "Montos"]:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1, f"Sheet '{sheet_name}' should have only header row, got {len(rows)}"


def test_build_excel_headers_are_bold(db_session):
    """Header row cells should have bold font."""
    data = build_excel(db_session, 2099)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Procesos"]
    for cell in ws[1]:
        assert cell.font.bold, f"Cell {cell.coordinate} should be bold"


# ---------------------------------------------------------------------------
# T3 — GET /export/excel endpoint tests
# ---------------------------------------------------------------------------

def test_get_excel_content_type(client, admin_headers, db_session):
    """200 + correct Excel content-type."""
    _make_proceso(db_session, "2026-001")
    resp = client.get("/export/excel?anno=2026", headers=admin_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_get_excel_content_disposition(client, admin_headers, db_session):
    """Content-Disposition contains correct filename."""
    _make_proceso(db_session, "2026-001")
    resp = client.get("/export/excel?anno=2026", headers=admin_headers)
    assert resp.status_code == 200
    cd = resp.headers.get("content-disposition", "")
    assert "adquisiciones_tic_2026.xlsx" in cd


def test_get_excel_response_is_valid_xlsx(client, admin_headers, db_session):
    """Response body can be parsed as a valid .xlsx with 3 sheets."""
    _make_proceso(db_session, "2026-001")
    resp = client.get("/export/excel?anno=2026", headers=admin_headers)
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Procesos" in wb.sheetnames
    assert "Etapas" in wb.sheetnames
    assert "Montos" in wb.sheetnames


def test_excel_without_anno_returns_422(client, admin_headers):
    """Missing anno param → 422 (FastAPI validation)."""
    resp = client.get("/export/excel", headers=admin_headers)
    assert resp.status_code == 422


def test_auth_excel_401(client):
    """No auth token → 401."""
    resp = client.get("/export/excel?anno=2026")
    assert resp.status_code == 401


# ---------------------------------------------------------------------------
# T2 — build_pdf tests (service-level)
# ---------------------------------------------------------------------------

def test_build_pdf_valid_bytes(db_session):
    """build_pdf returns bytes starting with %PDF-."""
    p = _make_proceso(db_session, "2026-001")
    _make_etapa(db_session, p.id, "E01", "COMPLETADO")

    data = build_pdf(db_session, p.id)

    assert isinstance(data, bytes)
    assert data[:4] == b"%PDF", f"Expected PDF header, got: {data[:8]!r}"


def test_build_pdf_404_not_found(db_session):
    """build_pdf raises 404 for nonexistent proceso_id."""
    from fastapi import HTTPException
    with pytest.raises(HTTPException) as exc_info:
        build_pdf(db_session, 99999)
    assert exc_info.value.status_code == 404


def test_build_pdf_404_soft_deleted(db_session):
    """build_pdf raises 404 for soft-deleted proceso."""
    from fastapi import HTTPException
    p = _make_proceso(
        db_session,
        "2026-soft",
        eliminado_en=datetime(2026, 1, 15, 10, 0, 0),
    )
    with pytest.raises(HTTPException) as exc_info:
        build_pdf(db_session, p.id)
    assert exc_info.value.status_code == 404


# ---------------------------------------------------------------------------
# T3 — GET /export/proceso/{id}/pdf endpoint tests
# ---------------------------------------------------------------------------

def test_get_pdf_endpoint(client, admin_headers, db_session):
    """200 + application/pdf + correct filename."""
    p = _make_proceso(db_session, "2026-001")
    _make_etapa(db_session, p.id, "E01", "COMPLETADO")

    resp = client.get(f"/export/proceso/{p.id}/pdf", headers=admin_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/pdf")
    cd = resp.headers.get("content-disposition", "")
    assert "proceso_2026-001.pdf" in cd


def test_get_pdf_starts_with_pdf_magic(client, admin_headers, db_session):
    """PDF response body starts with %PDF."""
    p = _make_proceso(db_session, "2026-001")
    resp = client.get(f"/export/proceso/{p.id}/pdf", headers=admin_headers)
    assert resp.status_code == 200
    assert resp.content[:4] == b"%PDF"


def test_get_pdf_404_not_found(client, admin_headers):
    """Nonexistent proceso_id → 404."""
    resp = client.get("/export/proceso/99999/pdf", headers=admin_headers)
    assert resp.status_code == 404
    assert resp.json()["detail"] == "Proceso no encontrado"


def test_get_pdf_404_soft_deleted(client, admin_headers, db_session):
    """Soft-deleted proceso → 404."""
    p = _make_proceso(
        db_session,
        "2026-soft",
        eliminado_en=datetime(2026, 1, 15, 10, 0, 0),
    )
    resp = client.get(f"/export/proceso/{p.id}/pdf", headers=admin_headers)
    assert resp.status_code == 404


def test_auth_pdf_401(client, db_session):
    """No auth token → 401."""
    p = _make_proceso(db_session, "2026-001")
    resp = client.get(f"/export/proceso/{p.id}/pdf")
    assert resp.status_code == 401


# ---------------------------------------------------------------------------
# T4 — Soft-deleted exclusion from Excel
# ---------------------------------------------------------------------------

def test_soft_deleted_excluded_from_excel(client, admin_headers, db_session):
    """Soft-deleted proceso does NOT appear in Procesos sheet."""
    _make_proceso(db_session, "2026-001")  # visible
    _make_proceso(
        db_session,
        "2026-soft",
        eliminado_en=datetime(2026, 1, 15, 10, 0, 0),
    )  # soft-deleted

    resp = client.get("/export/excel?anno=2026", headers=admin_headers)
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Procesos"]
    rows = list(ws.iter_rows(values_only=True))
    # Only 1 data row (header + 1 visible proceso)
    assert len(rows) == 2, f"Expected 2 rows (header+1), got {len(rows)}"
    # The soft-deleted id should not appear
    id_values = [r[0] for r in rows[1:]]
    assert "2026-soft" not in id_values
